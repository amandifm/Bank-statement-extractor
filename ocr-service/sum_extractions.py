import os
import glob
from openpyxl import load_workbook

extracted_dir = "C:/Users/amank/Downloads/Office/DIFM-Extractor/Extracted files"
excel_files = glob.glob(f"{extracted_dir}/*.xlsx")

for excel_file in excel_files:
    try:
        wb = load_workbook(excel_file, data_only=True)
        if "Transactions" not in wb.sheetnames:
            continue
            
        sheet = wb["Transactions"]
        
        # Find header row
        header_row = -1
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if row and row[0] == "Date":
                header_row = i
                break
                
        if header_row == -1:
            continue
            
        total_debits = 0.0
        total_credits = 0.0
        
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i <= header_row:
                continue
            if row[0] is None and row[1] is None:
                continue
            if row[0] in ["Debits / Withdrawals", "Credits / Deposits"]:
                continue
                
            debit_str = str(row[2]) if row[2] is not None else "0"
            debit_str = debit_str.replace("$", "").replace(",", "")
            try:
                debit_amt = float(debit_str)
            except ValueError:
                debit_amt = 0.0
                
            credit_str = str(row[3]) if row[3] is not None else "0"
            credit_str = credit_str.replace("$", "").replace(",", "")
            try:
                credit_amt = float(credit_str)
            except ValueError:
                credit_amt = 0.0
                
            total_debits += debit_amt
            total_credits += credit_amt
                
        print(f"{os.path.basename(excel_file):<40} Debits: ${total_debits:10.2f} | Credits: ${total_credits:10.2f}")
        
    except Exception as e:
        print(f"Error processing {os.path.basename(excel_file)}: {e}")

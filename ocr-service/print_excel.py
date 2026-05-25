import os
import glob
from openpyxl import load_workbook

extracted_dir = "C:/Users/amank/Downloads/Office/DIFM-Extractor/Extracted files"
excel_files = glob.glob(f"{extracted_dir}/*.xlsx")

for excel_file in excel_files[:1]:
    wb = load_workbook(excel_file, data_only=True)
    sheet = wb["Transactions"]
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        print(row)
        if i > 5:
            break

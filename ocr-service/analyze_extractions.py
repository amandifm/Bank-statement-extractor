import os
import glob
from openpyxl import load_workbook

extracted_dir = "C:/Users/amank/Downloads/Office/DIFM-Extractor/Extracted files"
excel_files = glob.glob(f"{extracted_dir}/*.xlsx")

print(f"Found {len(excel_files)} Excel reports.")

for excel_file in excel_files:
    try:
        wb = load_workbook(excel_file, data_only=True)
        if "Transactions" not in wb.sheetnames:
            print(f"Report: {os.path.basename(excel_file)} has no Transactions sheet.")
            continue
            
        sheet = wb["Transactions"]
        
        # Find header row
        header_row = -1
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if row and row[0] == "Date":
                header_row = i
                break
                
        if header_row == -1:
            print(f"Report: {os.path.basename(excel_file)} has no Date header.")
            continue
            
        debits = 0
        credits = 0
        missing_dates = 0
        missing_desc = 0
        total_txns = 0
        
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i <= header_row:
                continue
            if row[0] is None and row[1] is None:
                continue
            if row[0] in ["Debits / Withdrawals", "Credits / Deposits"]:
                continue
                
            total_txns += 1
            if row[5] == "Debit": debits += 1
            if row[5] == "Credit": credits += 1
            if not row[0]: missing_dates += 1
            if not row[1]: missing_desc += 1
            
        print(f"\nReport: {os.path.basename(excel_file)}")
        print(f"  Total Rows: {total_txns}")
        print(f"  Debits: {debits}, Credits: {credits}")
        if missing_dates > 0: print(f"  WARNING: {missing_dates} missing dates.")
        if missing_desc > 0: print(f"  WARNING: {missing_desc} missing desc.")
        
    except Exception as e:
        print(f"Error processing {os.path.basename(excel_file)}: {e}")
